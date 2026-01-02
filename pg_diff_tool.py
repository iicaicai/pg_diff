import argparse
import psycopg2
import psycopg2.pool
import csv
import json
import os
import subprocess
import sys
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Default configuration based on user input
DEFAULT_DB_URL = os.getenv("PG_DB_URL", "postgresql://127.0.0.1:5432")
DEFAULT_DB_NAME = "logto"
SNAPSHOT_FILE = "migration_snapshot.json"

class DataVerifier:
    def __init__(self, db_url, db_name, threads=1, container_name=None, db_user="postgres", use_local=False):
        self.db_url = db_url
        self.db_name = db_name
        self.threads = threads
        self.container_name = container_name
        self.db_user = db_user
        self.use_local = use_local
        self.pool = None
        self._init_pool()

    def _init_pool(self):
        try:
            # Construct full DSN.
            dsn = self.db_url
            
            # If dsn doesn't specify dbname, append it
            # Simple heuristic: if it looks like a URI and doesn't end with /dbname
            if "://" in dsn:
                 if self.db_name and not dsn.rstrip('/').endswith(f"/{self.db_name}"):
                      dsn = f"{dsn.rstrip('/')}/{self.db_name}"
            # Else assume it's a libpq string or user handles it. 
            # We trust psycopg2 to handle various formats (URI or keyword=value)

            print(f"Initializing connection pool ({self.threads} connections) to: {self.db_name}...")
            # Pass dbname explicitely to override/ensure it in case DSN lacks it (e.g. "host=... user=...")
            self.pool = psycopg2.pool.ThreadedConnectionPool(
                minconn=1,
                maxconn=self.threads,
                dsn=dsn,
                dbname=self.db_name
            )
            print("Connection pool initialized.")
        except Exception as e:
            print(f"Error initializing connection pool: {e}")
            sys.exit(1)

    def close_pool(self):
        if self.pool:
            self.pool.closeall()

    def get_primary_key_column(self, conn, schema, table):
        """
        Finds the primary key column name for a given table using a specific connection.
        """
        query = """
            SELECT kcu.column_name
            FROM information_schema.table_constraints tco
            JOIN information_schema.key_column_usage kcu 
                 ON kcu.constraint_name = tco.constraint_name
                 AND kcu.constraint_schema = tco.constraint_schema
            WHERE tco.constraint_type = 'PRIMARY KEY'
            AND kcu.table_schema = %s
            AND kcu.table_name = %s
            LIMIT 1;
        """
        with conn.cursor() as cursor:
            cursor.execute(query, (schema, table))
            result = cursor.fetchone()
            return result[0] if result else None

    def process_table(self, schema, table):
        """
        Worker function to process a single table.
        """
        full_name = f'"{schema}"."{table}"'
        key = f"{schema}.{table}"
        result_data = {
            "count": 0,
            "checksum": 0,
            "pks": [],
            "pk_col": None,
            "error": None
        }

        conn = None
        try:
            conn = self.pool.getconn()
            with conn.cursor() as cursor:
                # Get Count and Checksum (Content Hash)
                # Checksum logic: Sum of 64-bit prefixes of row MD5s. Order-independent.
                # Cast to bigint to avoid overflow issues with simple integer, though sum might still be large.
                # Actually, bigint sum wrap-around is fine for checksum purposes, but PG sum() on bigint might error on overflow?
                # PG sum(bigint) returns numeric, so no overflow. We can store it as string or large int.
                query_stats = f"""
                    SELECT 
                        COUNT(*),
                        COALESCE(SUM(('x' || substr(md5(t::text), 1, 16))::bit(64)::bigint), 0)
                    FROM {full_name} t
                """
                cursor.execute(query_stats)
                row = cursor.fetchone()
                result_data["count"] = row[0]
                result_data["checksum"] = str(row[1]) # Store as string to preserve precision in JSON

                # Get PKs
                pk_col = self.get_primary_key_column(conn, schema, table)
                result_data["pk_col"] = pk_col

                if pk_col:
                    # Optimized for large datasets: Only fetch PKs if count is reasonable?
                    # For now, we fetch all as requested, but user should be aware of memory.
                    cursor.execute(f"SELECT \"{pk_col}\" FROM {full_name}")
                    pks = [str(row[0]) for row in cursor.fetchall()]
                    pks.sort()
                    result_data["pks"] = pks
        
        except Exception as e:
            print(f"Error processing {full_name}: {e}")
            result_data["error"] = str(e)
        finally:
            if conn:
                self.pool.putconn(conn)
        
        return key, result_data

    def get_table_snapshot_data(self):
        """
        Retrieves row counts and primary keys for all tables using multithreading.
        """
        # Get list of tables first using a temporary connection
        conn = self.pool.getconn()
        try:
            with conn.cursor() as cursor:
                query_tables = """
                    SELECT n.nspname AS schema_name, c.relname AS table_name
                    FROM pg_class c
                    JOIN pg_namespace n ON n.oid = c.relnamespace
                    WHERE c.relkind = 'r' 
                    AND n.nspname NOT IN ('pg_catalog', 'information_schema')
                    ORDER BY n.nspname, c.relname;
                """
                cursor.execute(query_tables)
                tables = cursor.fetchall()
        finally:
            self.pool.putconn(conn)
        
        print(f"Scanning {len(tables)} tables with {self.threads} threads...")
        snapshot = {}
        
        with ThreadPoolExecutor(max_workers=self.threads) as executor:
            futures = [executor.submit(self.process_table, schema, table) for schema, table in tables]
            
            for future in as_completed(futures):
                key, data = future.result()
                snapshot[key] = data
        
        return snapshot

    def perform_backup(self, output_file=None):
        """
        Executes pg_dump via docker exec or locally.
        """
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"{self.db_name}_backup_{timestamp}.sql"
        
        print(f"Starting backup for {self.db_name} to {output_file}...")
        
        cmd = []
        if self.use_local:
             # Local pg_dump
             # Construct connection string for pg_dump
             conn_str = self.db_url
             if "://" in conn_str:
                 if self.db_name and not conn_str.rstrip('/').endswith(f"/{self.db_name}"):
                      conn_str = f"{conn_str.rstrip('/')}/{self.db_name}"
             else:
                 # KV style: append dbname
                 conn_str = f"{conn_str} dbname={self.db_name}"
             
             cmd = ["pg_dump", "--dbname", conn_str]
        else:
             # Docker exec
             if not self.container_name:
                 print("Error: Container name is required for Docker mode. Use --container-name or switch to --local.")
                 sys.exit(1)
                 
             cmd = [
                "docker", "exec", "-i", self.container_name, 
                "pg_dump", "-U", self.db_user, self.db_name
            ]
        
        try:
            with open(output_file, 'w') as outfile:
                subprocess.check_call(cmd, stdout=outfile)
            print(f"Backup completed successfully: {output_file}")
        except subprocess.CalledProcessError as e:
            print(f"Error running backup: {e}")
            if not self.use_local:
                 print("Ensure the container is running and accessible.")
        except Exception as e:
            print(f"An error occurred during backup: {e}")

def save_snapshot(data, filename):
    with open(filename, 'w') as f:
        json.dump(data, f, indent=2)
    print(f"Snapshot saved to {filename}")

def load_snapshot(filename):
    if not os.path.exists(filename):
        print(f"Error: Snapshot file {filename} not found.")
        sys.exit(1)
    with open(filename, 'r') as f:
        return json.load(f)

def generate_excel_report(before_snapshot, after_snapshot, output_file):
    """
    Compares before and after snapshots and writes Excel report with multiple sheets.
    """
    wb = Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(['Schema', 'Table', 'Before Count', 'After Count', 'Is Change', 'Change Type'])
    
    # Style header
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Sheet 2: Diff Details
    ws_details = wb.create_sheet(title="Diff Details")
    ws_details.append(['Schema', 'Table', 'Change Type', 'IDs/Details'])
    for cell in ws_details[1]:
        cell.font = header_font
        cell.fill = header_fill

    all_keys = set(before_snapshot.keys()) | set(after_snapshot.keys())
    
    print("Generating report...")
    for key in sorted(all_keys):
        schema, table = key.split('.', 1)
        
        # Helper
        def get_data(snapshot, k):
            val = snapshot.get(k, {'count': 0, 'checksum': '0', 'pks': []})
            if isinstance(val, int): 
                return {'count': val, 'checksum': '0', 'pks': []}
            return val

        data_before = get_data(before_snapshot, key)
        data_after = get_data(after_snapshot, key)
        
        count_before = data_before['count']
        count_after = data_after['count']
        
        checksum_before = str(data_before.get('checksum', '0'))
        checksum_after = str(data_after.get('checksum', '0'))
        
        pks_before = set(data_before.get('pks', []))
        pks_after = set(data_after.get('pks', []))
        
        # Diff calculation
        missing_pks = pks_before - pks_after
        added_pks = pks_after - pks_before
        
        is_change = 'N'
        change_type = []
        
        if count_before != count_after:
            change_type.append("Count Mismatch")
        
        if missing_pks or added_pks:
             # Already covered by ID check, but just to be sure
             pass
             
        if count_before == count_after and checksum_before != checksum_after and not missing_pks and not added_pks:
            change_type.append("Content Mismatch")
        
        if count_before != count_after or missing_pks or added_pks or checksum_before != checksum_after:
            is_change = 'Y'
            
            # Add to Details Sheet
            if missing_pks:
                missing_list = sorted(list(missing_pks))
                # Excel has cell limit (32k chars), careful with huge lists
                msg = ", ".join(missing_list)
                if len(msg) > 32000: msg = msg[:32000] + "...(truncated)"
                ws_details.append([schema, table, 'Missing IDs', msg])
                change_type.append("Missing IDs")
                
            if added_pks:
                added_list = sorted(list(added_pks))
                msg = ", ".join(added_list)
                if len(msg) > 32000: msg = msg[:32000] + "...(truncated)"
                ws_details.append([schema, table, 'Added IDs', msg])
                change_type.append("Added IDs")
            
            if not missing_pks and not added_pks and count_before != count_after:
                 ws_details.append([schema, table, 'Count Mismatch', 'No PK differences found (possibly duplicate PKs or no PK)'])
                 
            if not missing_pks and not added_pks and count_before == count_after and checksum_before != checksum_after:
                 ws_details.append([schema, table, 'Content Mismatch', f'Row count identical ({count_before}), but content checksum differs.'])

        ws_summary.append([schema, table, count_before, count_after, is_change, ", ".join(set(change_type))])

    wb.save(output_file)
    print(f"Comparison report generated: {output_file}")

def main():
    parser = argparse.ArgumentParser(description="PostgreSQL Upgrade Verification Tool")
    subparsers = parser.add_subparsers(dest='command', required=True)
    
    # Common arguments
    parent_parser = argparse.ArgumentParser(add_help=False)
    parent_parser.add_argument('--db-url', default=DEFAULT_DB_URL, help="Database connection string (base URL) or DSN")
    parent_parser.add_argument('--db-name', required=True, help="Database name to process")
    parent_parser.add_argument('--threads', type=int, default=4, help="Number of threads for parallel processing (default: 4)")
    
    # Mode selection
    parent_parser.add_argument('--container-name', default="docker-tmp-postgres-1", help="Docker container name (default: docker-tmp-postgres-1)")
    parent_parser.add_argument('--local', action='store_true', help="Use local pg_dump instead of Docker exec")
    parent_parser.add_argument('--db-user', default="postgres", help="Database user for pg_dump (default: postgres)")

    # Backup/Pre-upgrade command
    parser_backup = subparsers.add_parser('backup', parents=[parent_parser], help="Backup data and take snapshot of row counts")
    parser_backup.add_argument('--snapshot-file', default=SNAPSHOT_FILE, help="File to save row counts")
    parser_backup.add_argument('--dump-file', help="File to save SQL dump")
    parser_backup.add_argument('--skip-dump', action='store_true', help="Skip physical backup, only count rows")

    # Compare/Post-upgrade command
    parser_compare = subparsers.add_parser('compare', parents=[parent_parser], help="Compare current row counts with snapshot")
    parser_compare.add_argument('--snapshot-file', default=SNAPSHOT_FILE, help="Snapshot file to compare against")
    parser_compare.add_argument('--output', default='upgrade_diff_report.xlsx', help="Output Excel file for report")

    args = parser.parse_args()
    
    # If using local mode, container_name is irrelevant but we can leave it as None or ignored.
    # If container name is default but user might want to clear it? No, arg default handles it.
    
    verifier = DataVerifier(
        db_url=args.db_url, 
        db_name=args.db_name, 
        threads=args.threads,
        container_name=args.container_name,
        db_user=args.db_user,
        use_local=args.local
    )
    
    if args.command == 'backup':
        # 1. Physical Backup
        if not args.skip_dump:
            verifier.perform_backup(args.dump_file)
        
        # 2. Logical Snapshot (Counting)
        print("Taking pre-upgrade snapshot (with PKs)...")
        snapshot_data = verifier.get_table_snapshot_data()
        save_snapshot(snapshot_data, args.snapshot_file)
        
    elif args.command == 'compare':
        # 1. Load previous snapshot
        print("Loading previous snapshot...")
        before_snapshot = load_snapshot(args.snapshot_file)
        
        # 2. Get current counts
        print("Taking post-upgrade snapshot...")
        after_snapshot = verifier.get_table_snapshot_data()
        
        # 3. Generate Report
        generate_excel_report(before_snapshot, after_snapshot, args.output)
    
    verifier.close_pool()

if __name__ == "__main__":
    main()
